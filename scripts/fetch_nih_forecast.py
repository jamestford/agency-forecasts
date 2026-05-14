#!/usr/bin/env python3
"""Download the NIH (National Institutes of Health) procurement forecast.

NIH OAMP publishes quarterly forecast snapshots at predictable URLs:
  /sites/default/files/SBODocs/NIH <Nth> Quarter Procurement Forecast.xlsx

We try Q4 → Q1 in order and use the highest-quarter file that's
available. NIH's OAMP server has been observed to be flaky (CF 522);
we retry transient failures.

Schema: 'Contract' sheet, header row 1, ~165 real data rows padded
with ~52k blank rows we filter out. No clean primary key; we use a
composite (Contracting Office + Contract Name + Contract Description),
~96% unique.
"""
from __future__ import annotations

import datetime as _dt
import hashlib
import json
import sys
import time
import urllib.parse
from pathlib import Path

import openpyxl
import requests

sys.path.insert(0, str(Path(__file__).resolve().parent))
import agency_common as common
import diff_lib

AGENCY = "nih"
PAGE_URL = "https://oamp.od.nih.gov/division-of-acquisition-policy-and-evaluation/acquisition-resources/documents"
BASE = "https://oamp.od.nih.gov/sites/default/files/SBODocs/"
QUARTER_FILES = [
    "NIH 4th Quarter Procurement Forecast.xlsx",
    "NIH 3rd Quarter Procurement Forecast.xlsx",
    "NIH 2nd Quarter Procurement Forecast.xlsx",
    "NIH 1st Quarter Procurement Forecast.xlsx",
]

HEADER_ROW = 1
SHEET_NAME = "Contract"
KEY_COMPOSITE = ("Contracting Office Name", "Contract Name", "Contract Description")
TITLE_COLUMN = "Contract Name"
PREVIEW_COLUMNS = [
    "Contract Name",
    "Contracting Office Name",
    "Target Award Date",
    "FiscalYearContractRange",
    "TotalContractRange",
    "NAICSCode",
]


def _normalize(v):
    if isinstance(v, _dt.datetime):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def parse_rows(path: Path) -> tuple[dict, list[str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, []
    headers = [
        str(h).strip() if h is not None else f"_col{i}"
        for i, h in enumerate(rows[HEADER_ROW - 1])
    ]
    try:
        title_idx = headers.index("Contract Name")
    except ValueError:
        return {}, headers
    out: dict = {}
    for row in rows[HEADER_ROW:]:
        if not row or row[title_idx] is None or not str(row[title_idx]).strip():
            continue
        row_dict = {h: _normalize(row[i]) if i < len(row) else None for i, h in enumerate(headers)}
        key_parts = tuple(str(row_dict.get(c, "") or "") for c in KEY_COMPOSITE)
        key = " | ".join(key_parts)
        if not key.strip(" |"):
            continue
        out[key] = row_dict
    return out, headers


def fetch_latest_quarter() -> tuple[bytes, str, dict]:
    """Try Q4 → Q1; return (body, filename, response_headers) for the first that 200s."""
    last_error = None
    for fname in QUARTER_FILES:
        url = BASE + urllib.parse.quote(fname)
        for attempt in range(3):
            try:
                r = requests.get(url, headers={"User-Agent": common.UA, "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}, timeout=60)
            except requests.RequestException as e:
                print(f"  {fname} attempt {attempt+1}: {e}", flush=True)
                time.sleep(3)
                continue
            if r.status_code == 200 and len(r.content) > 10000:
                print(f"  {fname}: HTTP 200, {len(r.content):,} bytes", flush=True)
                return r.content, fname, {k.lower(): v for k, v in r.headers.items()}
            else:
                print(f"  {fname} attempt {attempt+1}: HTTP {r.status_code} ({len(r.content)}b)", flush=True)
                last_error = f"{fname}: HTTP {r.status_code}"
                if r.status_code != 522:  # 522 = CF transient; retry. Anything else = real 404/etc, skip.
                    break
                time.sleep(3)
        else:
            continue
    raise RuntimeError(f"No NIH quarter file fetchable. Last error: {last_error}")


def main() -> int:
    print(f"[fetch] probing NIH quarter files…", flush=True)
    body, filename, headers = fetch_latest_quarter()

    changed, meta, diff_summary = common.process_download(
        agency=AGENCY,
        page_url=PAGE_URL,
        file_url=BASE + urllib.parse.quote(filename),
        body=body,
        headers=headers,
        extension=".xlsx",
        filename=filename,
        parse_rows=parse_rows,
        title_column=TITLE_COLUMN,
        preview_columns=PREVIEW_COLUMNS,
    )

    if changed:
        print(f"  -> CHANGED. diff: {diff_summary or '(initial)'}", flush=True)
    else:
        print(f"  -> unchanged (sha matches; last changed {meta['last_changed_utc']})", flush=True)

    today = _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%d")
    common.emit_output(
        changed="true" if changed else "false",
        file_count="1",
        filenames=filename,
        source_last_modified=headers.get("last-modified") or "",
        diff_summary=diff_summary,
        archive_date=today if changed else "",
    )
    print(f"[fetch] done. changed={changed}", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
